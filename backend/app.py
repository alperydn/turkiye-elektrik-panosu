"""
Türkiye Elektrik Panosu — EPİAŞ + TEİAŞ proxy backend (FastAPI)
------------------------------------------------------------------
- /api/uretim      : CANLI. EPİAŞ rt-gen'den saatlik, kaynak bazinda uretim (MW).
- /api/kurulu-guc  : TEİAŞ'in resmi Excel'inden (kurulu_guc_teias.xlsx) en son ayin
                     kaynak bazinda kurulu gucu.
- /api/tarihsel    : Ayni Excel'den 2001'den bugune yillik kurulu guc serisi.

TEİAŞ verisini guncellemek icin: YTBS'den yeni Excel'i indirip
"kurulu_guc_teias.xlsx" adiyla bu dosyanin uzerine yaz. Kod her istekte
(kisa bir onbellekle) dosyayi yeniden okur, baska bir sey yapmana gerek yok.

Kimlik (EPİAŞ): ayni klasorde .env dosyasi:
    EPTR_USERNAME=eposta@ornek.com
    EPTR_PASSWORD=sifre

Calistirma:
    pip install -r requirements.txt
    uvicorn app:app --host 0.0.0.0 --port 8000
"""

import os
import time
from datetime import date
from pathlib import Path

import openpyxl
import requests
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from eptr2 import EPTR2

# --- EPİAŞ kimlik: once .env, yoksa ortam degiskeni ---
_u = os.environ.get("EPTR_USERNAME")
_p = os.environ.get("EPTR_PASSWORD")
try:
    eptr = EPTR2(username=_u, password=_p) if (_u and _p) else EPTR2(use_dotenv=True)
except Exception as ex:
    raise RuntimeError(f"EPTR2 baslatilamadi. .env veya ortam degiskenlerini kontrol edin: {ex}")

ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "*").split(",")

app = FastAPI(title="Türkiye Elektrik Panosu API", version="3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS],
    allow_methods=["GET"],
    allow_headers=["*"],
)

# ------------------------------------------------------------------
#  TEİAŞ KURULU GUC EXCEL'I — okuma ve donusum
# ------------------------------------------------------------------
XLSX_PATH = Path(__file__).parent / "kurulu_guc_teias.xlsx"
URETIM_XLSX_PATH = Path(__file__).parent / "uretim_tarihsel_teias.xlsx"
HEDEF_XLSX_PATH = Path(__file__).parent / "hedefler_etkb.xlsx"

# Panonun kaynak anahtarlarina renk + yenilenebilir bilgisi
SOURCE_META = {
    "gunes":      {"name": "Güneş",            "color": "#FBBF24", "renewable": True},
    "dogalgaz":   {"name": "Doğal Gaz",        "color": "#F97316", "renewable": False},
    "barajli":    {"name": "Barajlı Hidro",    "color": "#3B82F6", "renewable": True},
    "ruzgar":     {"name": "Rüzgar",           "color": "#2DD4BF", "renewable": True},
    "linyit":     {"name": "Yerli Kömür",      "color": "#B45309", "renewable": False},
    "ithalKomur": {"name": "İthal Kömür",      "color": "#64748B", "renewable": False},
    "akarsu":     {"name": "Akarsu Hidro",     "color": "#38BDF8", "renewable": True},
    "biyokutle":  {"name": "Biyokütle",        "color": "#84CC16", "renewable": True},
    "jeotermal":  {"name": "Jeotermal",        "color": "#EC4899", "renewable": True},
    "diger":      {"name": "Fuel Oil / Diğer", "color": "#94A3B8", "renewable": False},
}


def _read_teias_rows():
    """Excel'deki 'Veri' sayfasini okuyup her satiri panonun kaynak
    anahtarlarina (gunes, dogalgaz, barajli, ...) esler."""
    if not XLSX_PATH.exists():
        raise HTTPException(
            status_code=500,
            detail=f"TEİAŞ Excel dosyasi bulunamadi: {XLSX_PATH.name}. "
                   f"Dosyayi backend klasorune koyun.",
        )
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Veri"]
    raw = list(ws.iter_rows(min_row=1, values_only=True))
    header = raw[0]
    idx = {name: i for i, name in enumerate(header)}

    def g(row, col, default=0.0):
        i = idx.get(col)
        v = row[i] if i is not None else None
        return float(v) if v is not None else default

    out = []
    for row in raw[1:]:
        period = row[0]  # "YYYY-MM"
        if not period:
            continue
        year, month = period.split("-")
        vals = {
            "gunes":      g(row, "Güneş (MW)"),
            "dogalgaz":   g(row, "Doğal Gaz (MW)") + g(row, "LNG (MW)"),
            "barajli":    g(row, "Barajlı (MW)"),
            "ruzgar":     g(row, "Rüzgar (MW)"),
            "linyit":     g(row, "Linyit (MW)") + g(row, "Taş Kömür (MW)") + g(row, "Asfaltit Kömür (MW)"),
            "ithalKomur": g(row, "İthal Kömür (MW)"),
            "akarsu":     g(row, "Akarsu (MW)"),
            "biyokutle":  g(row, "Biyokütle (MW)"),
            "jeotermal":  g(row, "Jeotermal (MW)"),
            "diger":      g(row, "Fuel Oil (MW)") + g(row, "Motorin (MW)") + g(row, "Nafta (MW)") + g(row, "Atık Isı (MW)"),
        }
        vals["toplam"] = g(row, "Toplam (MW)") or sum(vals.values())
        out.append({"year": year, "month": month, "period": period, **vals})
    return out


def _kurulu_guc_from_teias():
    rows = _read_teias_rows()
    last = rows[-1]  # en guncel ay
    kaynaklar = []
    for key, meta in SOURCE_META.items():
        kaynaklar.append({
            "key": key,
            "name": meta["name"],
            "mw": round(last[key]),
            "color": meta["color"],
            "renewable": meta["renewable"],
        })
    return {"as_of": last["period"], "kaynaklar": kaynaklar}


def _tarihsel_uretim_from_teias():
    """'GENEL_YILLIK_ISLETME_NETICESI' raporundan (kaynak bazinda aylik
    uretim, kWh/MWh karisik) yillik toplam uretimi GWh olarak hesaplar."""
    if not URETIM_XLSX_PATH.exists():
        raise HTTPException(
            status_code=500,
            detail=f"Uretim Excel dosyasi bulunamadi: {URETIM_XLSX_PATH.name}.",
        )
    wb = openpyxl.load_workbook(URETIM_XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idxs = [i for i, r in enumerate(rows) if r[1] == "AY"]
    KEYS = ["dogalgaz", "linyit", "ithalKomur", "diger", "biyokutle",
            "jeotermal", "akarsu", "barajli", "gunes", "ruzgar", "toplam"]
    yearly = {}
    month_count = {}
    for hi in header_idxs:
        unit = rows[hi][-1] or ""
        is_kwh = "kWh" in unit
        year = str(rows[hi - 1][2]).strip()
        j = hi + 1
        while j < len(rows) and rows[j][1] not in (None, "AY"):
            r = rows[j]
            v = lambda i: float(r[i]) if r[i] is not None else 0.0
            vals = {
                "dogalgaz":   v(2) + v(3),
                "linyit":     v(4) + v(5) + v(6),
                "ithalKomur": v(7),
                "diger":      v(8) + v(9) + v(10) + v(11) + v(12),
                "biyokutle":  v(13),
                "jeotermal":  v(15),
                "akarsu":     v(16),
                "barajli":    v(17),
                "gunes":      v(19),
                "ruzgar":     v(20),
                "toplam":     v(21),
            }
            if is_kwh:
                vals = {k: val / 1000 for k, val in vals.items()}  # kWh -> MWh
            yearly.setdefault(year, {k: 0.0 for k in KEYS})
            month_count[year] = month_count.get(year, 0) + 1
            for k in KEYS:
                yearly[year][k] += vals[k]
            j += 1

    out = []
    for year in sorted(yearly.keys()):
        row = {"year": year, "ay_sayisi": month_count[year]}
        for k in KEYS:
            row[k] = round(yearly[year][k] / 1000, 1)  # MWh -> GWh
        row["hidro"] = round(row["akarsu"] + row["barajli"], 1)  # HIST_SOURCES grafiginde kullanilan birlesik alan
        out.append(row)
    return out


def _tarihsel_from_teias():
    """TEİAŞ kurulu guc Excel'inden yillik kurulu guc serisi (2001-bugun)."""
    rows = _read_teias_rows()
    by_year = {}
    for r in rows:
        by_year[r["year"]] = r  # ayni yil icinde ustune yazar -> en son ay kalir (Aralik varsa Aralik)
    out = []
    for year in sorted(by_year.keys()):
        r = by_year[year]
        out.append({
            "year": year,
            "toplam": round(r["toplam"]),
            "gunes": round(r["gunes"]),
            "ruzgar": round(r["ruzgar"]),
            "hidro": round(r["barajli"] + r["akarsu"]),
            "dogalgaz": round(r["dogalgaz"]),
            "linyit": round(r["linyit"]),
            "ithalKomur": round(r["ithalKomur"]),
            "jeotermal": round(r["jeotermal"]),
            "biyokutle": round(r["biyokutle"]),
            "diger": round(r["diger"]),
        })
    return out


# ------------------------------------------------------------------
#  ETKB 2024-2028 STRATEJIK PLANI - PERFORMANS GOSTERGELERI
# ------------------------------------------------------------------
def _read_hedef_rows():
    """'hedefler_etkb.xlsx' -> Amac/Hedef/Gosterge hiyerarsisiyle satirlar."""
    if not HEDEF_XLSX_PATH.exists():
        raise HTTPException(
            status_code=500,
            detail=f"Hedefler Excel dosyasi bulunamadi: {HEDEF_XLSX_PATH.name}.",
        )
    wb = openpyxl.load_workbook(HEDEF_XLSX_PATH, data_only=True)
    ws = wb["Tüm Amaç ve Göstergeler"]
    rows = list(ws.iter_rows(values_only=True))

    amac_kodu = amac_adi = hedef_kodu = hedef_adi = None
    out = []
    for r in rows:
        if not r or r[0] is None:
            continue
        first = str(r[0]).strip()
        if first.startswith("Amaç"):
            amac_kodu, amac_adi = first, r[1]
        elif first.startswith("Hedef"):
            hedef_kodu, hedef_adi = first, r[1]
        elif first.startswith("PG "):
            out.append({
                "gosterge_kodu": first,
                "gosterge_adi": r[1],
                "agirlik": float(r[2]) if r[2] is not None else None,
                "baseline": float(r[3]) if r[3] is not None else None,
                "2024": float(r[4]) if r[4] is not None else None,
                "2025": float(r[5]) if r[5] is not None else None,
                "2026": float(r[6]) if r[6] is not None else None,
                "2027": float(r[7]) if r[7] is not None else None,
                "2028": float(r[8]) if r[8] is not None else None,
                "amac_kodu": amac_kodu, "amac_adi": amac_adi,
                "hedef_kodu": hedef_kodu, "hedef_adi": hedef_adi,
            })
    return out


_GOSTERGE_HESAP = {
    "PG 1.1.1": lambda kg, ur: kg.get("gunes", 0),
    "PG 1.1.2": lambda kg, ur: kg.get("ruzgar", 0),
    "PG 1.1.3": lambda kg, ur: kg.get("barajli", 0) + kg.get("akarsu", 0) + kg.get("jeotermal", 0) + kg.get("biyokutle", 0),
    "PG 1.1.4": lambda kg, ur: 0,  # nukleer kurulu guc - TEİAŞ verisinde ayri kolon yok, gerceklesen=0 kabul edilir
    "PG 1.1.5": lambda kg, ur: kg.get("dogalgaz", 0) + kg.get("linyit", 0) + kg.get("ithalKomur", 0) + kg.get("diger", 0),
    "PG 2.1.1": lambda kg, ur: round((ur["toplam"] - ur["dogalgaz"] - ur["ithalKomur"]) / 1000, 1),
    "PG 2.1.2": lambda kg, ur: round((ur["toplam"] - ur["dogalgaz"] - ur["ithalKomur"]) / ur["toplam"] * 100, 2),
    "PG 3.1.1": lambda kg, ur: round((ur["gunes"] + ur["ruzgar"] + ur["hidro"] + ur["jeotermal"] + ur["biyokutle"]) / ur["toplam"] * 100, 2),
    "PG 3.1.2": lambda kg, ur: round(ur["gunes"] / ur["toplam"] * 100, 2),
    "PG 3.1.3": lambda kg, ur: round(ur["ruzgar"] / ur["toplam"] * 100, 2),
    "PG 3.1.4": lambda kg, ur: 0,  # nukleer uretim payi - uretim verisinde ayri kolon yok, gerceklesen=0 kabul edilir
}


def _hedefler_tamamlanma():
    rows = _read_hedef_rows()
    kg_list = _kurulu_guc_from_teias()
    kg_map = {k["key"]: k["mw"] for k in kg_list["kaynaklar"]}
    kg_asof = kg_list["as_of"]

    ur_all = _tarihsel_uretim_from_teias()
    full_years = [r for r in ur_all if r.get("ay_sayisi") == 12]
    ur_row = full_years[-1] if full_years else (ur_all[-1] if ur_all else None)
    ur_year = ur_row["year"] if ur_row else None

    out = []
    for r in rows:
        code = r["gosterge_kodu"]
        is_capacity = code.startswith("PG 1.1")
        is_pct = "(%)" in (r["gosterge_adi"] or "")
        calc = _GOSTERGE_HESAP.get(code)

        deger = None
        if calc is not None and (kg_map if is_capacity else ur_row):
            try:
                deger = calc(kg_map, ur_row)
            except Exception:
                deger = None
        veri_var = deger is not None

        if is_capacity:
            donem, hedef_yil = kg_asof, "2026"
        else:
            donem = ur_year
            hedef_yil = ur_year if ur_year in ("2024", "2025", "2026", "2027", "2028") else "2025"

        baseline, hedef_deger, hedef_2028 = r.get("baseline"), r.get(hedef_yil), r.get("2028")
        if is_pct:
            baseline = round(baseline * 100, 2) if baseline is not None else None
            hedef_deger = round(hedef_deger * 100, 2) if hedef_deger is not None else None
            hedef_2028 = round(hedef_2028 * 100, 2) if hedef_2028 is not None else None

        tamamlanma = None
        if veri_var and baseline is not None and hedef_2028 is not None and hedef_2028 != baseline:
            tamamlanma = round((deger - baseline) / (hedef_2028 - baseline) * 100, 1)

        out.append({
            "gosterge_kodu": code,
            "gosterge_adi": r["gosterge_adi"],
            "amac_kodu": r["amac_kodu"], "amac_adi": r["amac_adi"],
            "hedef_kodu": r["hedef_kodu"], "hedef_metni": r["hedef_adi"],
            "agirlik": r["agirlik"],
            "birim": "%" if is_pct else ("milyar kWh/yıl" if code == "PG 2.1.1" else "MW"),
            "baseline": baseline,
            "donem": donem,
            "gerceklesen": round(deger, 2) if veri_var else None,
            "hedef_donem_yili": hedef_yil,
            "hedef_donem_degeri": hedef_deger,
            "hedef_2028": hedef_2028,
            "tamamlanma_yuzde": tamamlanma,
            "veri_var": veri_var,
        })

    by_hedef = {}
    for it in out:
        g = by_hedef.setdefault(it["hedef_kodu"], {"metin": it["hedef_metni"], "items": []})
        g["items"].append(it)
    hedef_ozet = []
    for hkod, info in by_hedef.items():
        var_olanlar = [it for it in info["items"] if it["tamamlanma_yuzde"] is not None]
        skor = None
        if var_olanlar:
            wsum = sum(it["agirlik"] for it in var_olanlar)
            if wsum:
                skor = round(sum(it["agirlik"] * it["tamamlanma_yuzde"] for it in var_olanlar) / wsum, 1)
        hedef_ozet.append({
            "hedef_kodu": hkod, "hedef_metni": info["metin"],
            "tamamlanma_yuzde": skor,
            "gosterge_sayisi": len(info["items"]),
            "veri_olan_sayisi": len(var_olanlar),
        })

    return {"gostergeler": out, "hedef_ozet": hedef_ozet}


# ------------------------------------------------------------------
#  EPDK ELEKTRIK URETIM LISANSI - SANTRAL HARITASI
# ------------------------------------------------------------------
EPDK_URETIM_URL = "https://apigateway.epdk.gov.tr/elektrikUretimLisansiSorgula"


def _fetch_epdk_uretim_lisanslari():
    try:
        resp = requests.get(
            EPDK_URETIM_URL,
            json={"lisansDurumu": ["ONAYLANDI"]},
            timeout=60,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"EPDK uretim lisanslari alinamadi: {e}")


def _santral_harita_verisi():
    """EPDK'daki onayli uretim lisanslarini il ve kaynak turune gore toplar."""
    data = _fetch_epdk_uretim_lisanslari()
    by_il = {}
    kaynak_turleri = set()
    for lisans in data:
        for tesis in (lisans.get("uretimTesisi") or []):
            adres = tesis.get("adres") or {}
            il = adres.get("tesis_il") or lisans.get("il")
            if not il:
                continue
            il = il.strip().upper()
            tur = (tesis.get("tesisTuru") or "DIGER").strip().upper()
            mw = tesis.get("kuruluGucMWe") or 0.0
            entry = by_il.setdefault(il, {"toplam": 0.0, "kaynaklar": {}, "tesis_sayisi": 0})
            entry["toplam"] += mw
            entry["kaynaklar"][tur] = entry["kaynaklar"].get(tur, 0.0) + mw
            entry["tesis_sayisi"] += 1
            kaynak_turleri.add(tur)

    for il, v in by_il.items():
        v["toplam"] = round(v["toplam"], 1)
        v["kaynaklar"] = {k: round(val, 1) for k, val in v["kaynaklar"].items()}

    return {
        "iller": by_il,
        "kaynak_turleri": sorted(kaynak_turleri),
        "toplam_tesis": sum(v["tesis_sayisi"] for v in by_il.values()),
    }


# ------------------------------------------------------------------
#  rt-gen kolonlarini panonun kaynaklarina eslestirme (canli uretim)
# ------------------------------------------------------------------
def _to_panel(row):
    g = lambda k: float(row.get(k) or 0)
    r = {
        "saat":       str(row.get("hour", ""))[:5],
        "gunes":      round(g("sun")),
        "dogalgaz":   round(g("naturalGas")),
        "barajli":    round(g("dammedHydro")),
        "akarsu":     round(g("river")),
        "ruzgar":     round(g("wind")),
        "linyit":     round(g("lignite") + g("blackCoal") + g("asphaltiteCoal")),
        "ithalKomur": round(g("importCoal")),
        "biyokutle":  round(g("biomass")),
        "jeotermal":  round(g("geothermal")),
        "diger":      round(g("fueloil") + g("naphta") + g("lng") + g("wasteheat")),
    }
    r["toplam"] = sum(v for k, v in r.items() if k != "saat")
    return r


def _fiyatlar(start, end):
    """PTF (mcp), SMF (smp) ve AOF (wap) servislerini saat bazinda birlestirir."""
    def call(service, valcol):
        try:
            res = eptr.call(service, start_date=start, end_date=end)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"EPİAŞ {service} basarisiz: {e}")
        rows = res.to_dict(orient="records") if hasattr(res, "to_dict") else res
        return {r["date"]: float(r[valcol]) for r in rows if r.get("date") is not None}

    ptf_map = call("mcp", "price")
    smf_map = call("smp", "systemMarginalPrice")
    aof_map = call("wap", "wap")

    dates = sorted(set(ptf_map) | set(smf_map) | set(aof_map))
    out = []
    for d in dates:
        out.append({
            "saat": d[11:16],
            "ptf": round(ptf_map.get(d, 0), 2),
            "smf": round(smf_map.get(d, 0), 2),
            "aof": round(aof_map.get(d, 0), 2),
        })
    return out


# --- basit onbellek ---
_cache = {}
def cached(key, ttl, producer):
    now = time.time()
    hit = _cache.get(key)
    if hit and now - hit[0] < ttl:
        return hit[1]
    val = producer()
    _cache[key] = (now, val)
    return val


def _rt_gen(start, end):
    try:
        res = eptr.call("rt-gen", start_date=start, end_date=end)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"EPİAŞ rt-gen basarisiz: {e}")
    rows = res.to_dict(orient="records") if hasattr(res, "to_dict") else res
    return [_to_panel(r) for r in rows]


# ------------------------------------------------------------------
#  UCLAR
# ------------------------------------------------------------------
@app.get("/api/kurulu-guc")
def kurulu_guc():
    # TEİAŞ Excel'i pek sik degismiyor (ayda bir); 6 saat onbellek yeterli.
    return cached("kurulu-guc", 6 * 3600, _kurulu_guc_from_teias)


@app.get("/api/tarihsel")
def tarihsel():
    return cached("tarihsel", 6 * 3600, _tarihsel_from_teias)


@app.get("/api/uretim-tarihsel")
def uretim_tarihsel():
    return cached("uretim-tarihsel", 6 * 3600, _tarihsel_uretim_from_teias)


@app.get("/api/uretim")
def uretim(start: str | None = Query(None), end: str | None = Query(None)):
    end = end or date.today().isoformat()
    start = start or end
    return cached(f"uretim:{start}:{end}", 900, lambda: _rt_gen(start, end))


@app.get("/api/fiyatlar")
def fiyatlar(start: str | None = Query(None), end: str | None = Query(None)):
    end = end or date.today().isoformat()
    start = start or end
    return cached(f"fiyatlar:{start}:{end}", 900, lambda: _fiyatlar(start, end))


@app.get("/api/hedefler")
def hedefler():
    return cached("hedefler", 6 * 3600, _hedefler_tamamlanma)


@app.get("/api/santral-harita")
def santral_harita():
    return cached("santral-harita", 24 * 3600, _santral_harita_verisi)


@app.get("/health")
def health():
    return {"ok": True}

